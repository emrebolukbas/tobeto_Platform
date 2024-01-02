from selenium import webdriver
from selenium.webdriver.common.by import By
from webdriver_manager.chrome import ChromeDriverManager
from time import sleep
from selenium.webdriver.support.wait import WebDriverWait #ilgili driverı bekleten yapı
from selenium.webdriver.support import expected_conditions as ec #beklenen koşullar
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.keys import Keys
import pytest
import openpyxl
from constants import globalConstants as c

class Test_tobetoPlatformLogin():
    def setup_method(self):
        self.driver = webdriver.Chrome()
        self.driver.get(c.BASE_URL)
        self.driver.maximize_window()

    def teardown_metahod(self):
        self.driver.quit()
    
    def getData():
        excel = openpyxl.load_workbook("data/invalidLogin.xlsx")
        sheet = excel["Sheet1"]
        rows = sheet.max_row
        data = []
        for i in range(2,rows+1):
            email = sheet.cell(i,1).value
            password = sheet.cell(i,2).value
            data.append((email,password))

        return data
    
    #1)Giriş yap alanı görüntülenebilir ve işlevselliği test edilecektir.
    def test_visibility_of_login_page(self):
        tobeto_Img = WebDriverWait(self.driver,2).until(ec.visibility_of_element_located((By.XPATH,c.LOGIN_PAGE_TOBETO_IMG)))
        email_input = WebDriverWait(self.driver,2).until(ec.visibility_of_element_located((By.XPATH,c.E_MAIL_XPATH)))
        password_input = WebDriverWait(self.driver,2).until(ec.visibility_of_element_located((By.XPATH,c.PASSWORD_XPATH)))
        loginButton = WebDriverWait(self.driver,2).until(ec.visibility_of_element_located((By.XPATH,c.LOGIN_BUTTON_XPATH)))
        sign_up = WebDriverWait(self.driver,2).until(ec.visibility_of_element_located((By.XPATH,c.SIGN_UP_TEXT)))

        assert tobeto_Img,email_input and password_input and loginButton and sign_up.text == "Henüz üye değil misin? Kayıt Ol"
        
    #2)Kullanıcının e-posta veya şifre bilgilerini boş girerek sisteme giriş yapabilmesi test edilecektir.
    @pytest.mark.parametrize("email_param, password_param", [("", ""),("test@test.com",""),("","testSifre")])
    def test_passing_empty(self,email_param,password_param):
        eMail = WebDriverWait(self.driver,2).until(ec.visibility_of_element_located((By.XPATH,c.E_MAIL_XPATH)))
        eMail.send_keys(email_param)
        password = WebDriverWait(self.driver,2).until(ec.visibility_of_element_located((By.XPATH,c.PASSWORD_XPATH)))
        password.send_keys(password_param)    
        loginButton = WebDriverWait(self.driver,2).until(ec.visibility_of_element_located((By.XPATH,c.LOGIN_BUTTON_XPATH)))
        loginButton.click()
        errorMessage = WebDriverWait(self.driver,2).until(ec.visibility_of_element_located((By.XPATH,c.MANDATORY_FIELD_XPATH)))
        assert errorMessage.text == "Doldurulması zorunlu alan*"
    
    #3Kullanıcının e-posta ve şifre bilgilerini yanlış girerek sisteme giriş yapması  test edilecektir.
    @pytest.mark.parametrize("email_param, password_param", getData())
    def test_invlalidLogin(self,email_param,password_param):
        eMail = WebDriverWait(self.driver,2).until(ec.visibility_of_element_located((By.XPATH,c.E_MAIL_XPATH)))
        eMail.send_keys(email_param)
        password = WebDriverWait(self.driver,2).until(ec.visibility_of_element_located((By.XPATH,c.PASSWORD_XPATH)))
        password.send_keys(password_param)
        loginButton = WebDriverWait(self.driver,2).until(ec.visibility_of_element_located((By.XPATH,c.LOGIN_BUTTON_XPATH)))
        loginButton.click()
        
        systemMessage = WebDriverWait(self.driver,2).until(ec.presence_of_element_located((By.XPATH,c.SYSTEM_MESSAGE)))
        #assert systemMessage.text == "• Geçersiz e-posta veya şifre."
        assert "Geçersiz e-posta veya şifre." in systemMessage.text
         
"""     def test_password_passing_empty(self):
        eMail = WebDriverWait(self.driver,2).until(ec.visibility_of_element_located((By.XPATH,c.E_MAIL_XPATH)))
        eMail.send_keys("test@test.com")
        password = WebDriverWait(self.driver,2).until(ec.visibility_of_element_located((By.XPATH,c.PASSWORD_XPATH)))
        password.send_keys("")    
        loginButton = WebDriverWait(self.driver,2).until(ec.visibility_of_element_located((By.XPATH,c.LOGIN_BUTTON_XPATH)))
        loginButton.click()
        errorMessage = WebDriverWait(self.driver,2).until(ec.visibility_of_element_located((By.XPATH,c.MANDATORY_FIELD_XPATH)))
        assert errorMessage.text == "Doldurulması zorunlu alan*"
     
    def test_email_passing_empty(self):
        eMail = WebDriverWait(self.driver,2).until(ec.visibility_of_element_located((By.XPATH,c.E_MAIL_XPATH)))
        eMail.send_keys("")
        password = WebDriverWait(self.driver,2).until(ec.visibility_of_element_located((By.XPATH,c.PASSWORD_XPATH)))
        password.send_keys("testSifre")    
        loginButton = WebDriverWait(self.driver,2).until(ec.visibility_of_element_located((By.XPATH,c.LOGIN_BUTTON_XPATH)))
        loginButton.click()
        errorMessage = WebDriverWait(self.driver,2).until(ec.visibility_of_element_located((By.XPATH,c.MANDATORY_FIELD_XPATH)))
        assert errorMessage.text == "Doldurulması zorunlu alan*"
 """



        